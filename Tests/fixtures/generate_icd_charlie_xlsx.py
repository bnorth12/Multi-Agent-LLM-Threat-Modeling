"""Script to generate icd_charlie_v1.xlsx multi-sheet ICD fixture.

Sheet 1 (Entities): system hierarchy - subsystems, components, functions
Sheet 2 (Interfaces): interfaces between entities
"""
import pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = pathlib.Path(__file__).parent / "inputs" / "icd" / "icd_charlie_v1.xlsx"

wb = openpyxl.Workbook()

# ── Sheet 1: Entities ────────────────────────────────────────────────────────
ws_entities = wb.active
ws_entities.title = "Entities"

entity_headers = ["entity_type", "id", "name", "description", "parent", "hardware", "software_modules"]
ws_entities.append(entity_headers)

entities = [
    # subsystems
    ["subsystem", "SS-EGW-01", "Encryption Gateway Subsystem",
     "Performs in-line encryption and decryption of all satellite link traffic",
     "Charlie Satellite Communications Terminal", "", ""],
    ["subsystem", "SS-TRF-01", "Traffic Shaping Subsystem",
     "Controls data flow rates and priority queuing for the satellite link",
     "Charlie Satellite Communications Terminal", "", ""],
    # components
    ["component", "C-BEM-01", "Bulk Encryption Module",
     "Applies AES-256-GCM encryption and decryption to satellite data frames",
     "SS-EGW-01", "hosted", "egw.aes_engine"],
    ["component", "C-KDM-01", "Key Distribution Module",
     "Receives keying material and distributes keys to the encryption module on schedule",
     "SS-EGW-01", "hosted", "egw.key_dist"],
    ["component", "C-TRF-01", "Traffic Shaper",
     "Manages queuing, prioritization, and rate limiting for satellite link data",
     "SS-TRF-01", "hosted", "trf.shaper|trf.queue_mgr"],
    # functions
    ["function", "F-BEM-01", "Encrypt Outbound Frame",
     "Accepts a plaintext frame from the Traffic Shaper and applies AES-256-GCM encryption",
     "C-BEM-01", "", ""],
    ["function", "F-BEM-02", "Decrypt Inbound Frame",
     "Receives a ciphertext frame and decrypts it before forwarding to the Operations Network Interface",
     "C-BEM-01", "", ""],
    ["function", "F-KDM-01", "Receive Key Material",
     "Accepts new keying material from the Key Management Authority over the Key Distribution Interface",
     "C-KDM-01", "", ""],
    ["function", "F-KDM-02", "Activate Key",
     "Activates the next scheduled key and notifies the Bulk Encryption Module",
     "C-KDM-01", "", ""],
    ["function", "F-TRF-01", "Queue Data Frame",
     "Assigns a priority class to an incoming frame and places it in the transmit queue",
     "C-TRF-01", "", ""],
    ["function", "F-TRF-02", "Shape Transmit Rate",
     "Regulates outbound transmit rate and forwards frames to the Encryption Gateway",
     "C-TRF-01", "", ""],
]

for row in entities:
    ws_entities.append(row)

# Style header row
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_entities[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ws_entities.column_dimensions["A"].width = 14
ws_entities.column_dimensions["B"].width = 14
ws_entities.column_dimensions["C"].width = 32
ws_entities.column_dimensions["D"].width = 60
ws_entities.column_dimensions["E"].width = 40
ws_entities.column_dimensions["F"].width = 12
ws_entities.column_dimensions["G"].width = 28

# ── Sheet 2: Interfaces ───────────────────────────────────────────────────────
ws_ifaces = wb.create_sheet("Interfaces")

iface_headers = [
    "id", "name", "description",
    "from_node", "to_node", "interface_type",
    "protocol", "data_items",
    "trust_boundary_crossing", "trust_boundary_name"
]
ws_ifaces.append(iface_headers)

interfaces = [
    ["IF-201", "Plaintext Frame to Encryption Interface",
     "Carries plaintext frames from the Traffic Shaper function to the Bulk Encryption Module",
     "F-TRF-02", "F-BEM-01", "function-function",
     "internal", "plaintext_frame", "false", ""],
    ["IF-202", "Encrypted Frame to Satellite Modem Interface",
     "Carries AES-256-GCM encrypted frames from the Bulk Encryption Module to the Satellite Modem",
     "C-BEM-01", "SAT_MODEM", "component-external",
     "RF-link", "encrypted_frame", "true", "Satellite Link Boundary"],
    ["IF-203", "Inbound Satellite Frame Interface",
     "Carries inbound ciphertext frames from the Satellite Modem to the Bulk Encryption Module",
     "SAT_MODEM", "C-BEM-01", "external-component",
     "RF-link", "encrypted_frame", "true", "Satellite Link Boundary"],
    ["IF-204", "Key Distribution Interface",
     "Carries keying material from the Key Management Authority to the Key Distribution Module",
     "KEY_MGMT_AUTHORITY", "C-KDM-01", "external-component",
     "HTTPS", "key_material|key_id", "true", "Key Management Authority Boundary"],
    ["IF-205", "Key Activation Notification Interface",
     "Notifies the Bulk Encryption Module when a new key is active",
     "F-KDM-02", "C-BEM-01", "function-component",
     "internal", "key_activation_event", "false", ""],
    ["IF-206", "Operations Network Ingress Interface",
     "Carries data frames from the Operations Network to the Traffic Shaping Subsystem",
     "OPS_NETWORK", "SS-TRF-01", "external-subsystem",
     "HTTPS", "data_frame|source_auth_token", "true", "Operations Network Boundary"],
    ["IF-207", "Encryption to Operations Network Egress Interface",
     "Carries decrypted inbound frames from the Bulk Encryption Module to the Operations Network",
     "C-BEM-01", "OPS_NETWORK", "component-external",
     "HTTPS", "plaintext_frame", "true", "Operations Network Boundary"],
    ["IF-208", "Traffic Shaper Subsystem Internal Bus",
     "Internal bus between the Traffic Shaping Subsystem and the Encryption Gateway Subsystem",
     "SS-TRF-01", "SS-EGW-01", "subsystem-subsystem",
     "internal", "queued_frame", "false", ""],
]

for row in interfaces:
    ws_ifaces.append(row)

for cell in ws_ifaces[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ws_ifaces.column_dimensions["A"].width = 10
ws_ifaces.column_dimensions["B"].width = 40
ws_ifaces.column_dimensions["C"].width = 60
ws_ifaces.column_dimensions["D"].width = 22
ws_ifaces.column_dimensions["E"].width = 22
ws_ifaces.column_dimensions["F"].width = 22
ws_ifaces.column_dimensions["G"].width = 14
ws_ifaces.column_dimensions["H"].width = 30
ws_ifaces.column_dimensions["I"].width = 24
ws_ifaces.column_dimensions["J"].width = 34

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print(f"Wrote {OUT}")
